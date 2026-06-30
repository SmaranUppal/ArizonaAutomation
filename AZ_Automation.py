"""
AZ Courts — JG Wentworth Case Scraper
======================================
Searches the Arizona Courts public access portal for each business name variant,
clicks into cases filed in the last 60 days, and appends results to an Excel file.
Re-solves the CAPTCHA automatically whenever a verification screen reappears,
then resumes exactly where it left off.

USAGE:
  python az_courts_jgw.py
  python az_courts_jgw.py --out my_cases.xlsx
  python az_courts_jgw.py --days 90 --headless

REQUIREMENTS (same directory):
  captcha_auto.py  captcha_model.pth  captcha_chars.txt
  pip install selenium openpyxl requests opencv-python torch
"""

import argparse
import re
import time
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    NoSuchElementException,
    TimeoutException,
    StaleElementReferenceException,
)

try:
    from captcha_auto import load_model, predict, download_image
    CAPTCHA_AUTO = True
except ImportError:
    CAPTCHA_AUTO = False
    print("WARNING: captcha_auto.py not found — CAPTCHAs must be solved manually.")

# ── Constants ─────────────────────────────────────────────────────────────────

BASE_URL  = "https://apps.azcourts.gov/publicaccess/"
SEARCH_URL = BASE_URL + "caselookup.aspx?AspxAutoDetectCookieSupport=1"

BUSINESS_NAMES = [
    "JG Wentworth",
    "J G Wentworth",
    "J.G. Wentworth",
    "J. G. Wentworth",
    "Peachtree Settlement",
    "Peach Tree Settlement",
    "DRB Cap",
    "Stone Street",
    "AA Ron I",
    "Abactor",
    "Abidole",
    "Adenna Med",
    "Adventura",
    "AGPI",
    "Aikman Structured Finance",
    "Annuity Transfers Ltd",
    "Apis Management",
    "Atlas Legal Funding III LP",
    "AXE Finance",
    "B.A.W.21",
    "B.R. Wright",
    "BHG Structured",
    "Bifco",
    "Blue Grape",
    "Catalina Structured Funding",
    "Concordis Group Limited",
    "Conrad Factoring",
    "Cornerstone Funding",
    "Fast Annuity S",
    "FL Assignments Corp",
    "G.D.T.R.F.B.",
    "G7 Crescenta",
    "Genex Capital Corp",
    "GJ 123",
    "Greenwood Funding",
    "Grier I",
    "Hakstol Group",
    "Hiddenview Ent, LLC",
    "JLC Capital Funding",
    "KN Direct Capital",
    "Lane Nimitz",
    "Lasko LLC",
    "Lasko, LLC",
    "Leaf 002 LLC",
    "Legere LLC",
    "Legere, LLC",
    "Lottery Funding",
    "M McDougall LLC",
    "M McDougall, LLC",
    "Majestic Funding",
    "Mic-Bry8",
    "Olive Branch Funding",
    "Palermo Group",
    "Palm Green Closing",
    "Palm Harbor",
    "Passira Mal",
    "Patriot Settlement",
    "QLS Funding",
    "Reliance Funding",
    "Rocorp Corporation",
    "RSL Funding",
    "Savannah Settlements",
    "Sempra Finance",
    "Seneca Originations",
    "SeneOne LLC",
    "Settlement Capital Corp",
    "Settlement Status",
    "Somerton LLC",
    "Somerton, LLC",
    "Stratcap Investments",
    "Stratton Asset",
    "Structured Asset",
    "TKD LLC",
    "TKD, LLC",
    "TRM V LLC",
    "TRM V, LLC",
    "Tybenz LLC",
    "Tybenz, LLC",
    "Uber Funding",
    "Vintage Equity Group",
    "Wepaymore Funding",
    "Zakho Way",
    "GREAT PLAINS MANAGEMENT CORPORATION",
    "T ENE LLC",
    "T ENE, LLC",
    "RD FITZ LLC",
    "RD FITZ, LLC",
    "GA OFF LLC",
    "GA OFF, LLC",
    "Assured Management Corporation",
    "BENTZEN FINANCIAL",
    "Bentzen Funding",
]

CAPTCHA_IMG_ID = "caselookup_ctl00_contentplaceholder1_samplecaptcha_CaptchaImage"
CAPTCHA_INPUT  = "ctl00$ContentPlaceHolder1$CaptchaCodeTextBox"
CAPTCHA_SUBMIT = "ctl00$ContentPlaceHolder1$btnCaptcha"
MAX_CAPTCHA_TRIES = 30

SORT_JS       = "__doPostBack('ctl00$ContentPlaceHolder1$gvSearchResults','Sort$cs_srch_key')"
PAGE_JS       = "ctl00$ContentPlaceHolder1$gvSearchResults"  # prefix for Page$N postback

COLUMNS       = ["Case Number", "Case Title", "Court", "Filing Date", "Search Name"]
HEADER_COLOR  = "1F4E79"
ALT_ROW_COLOR = "D6E4F0"


# ── Helpers: year / date ──────────────────────────────────────────────────────

def extract_year_from_case(case_num: str) -> int | None:
    """Pull the 4-digit year from a case number like S-0400-CV-202600131."""
    for part in reversed(case_num.strip().split("-")):
        m = re.match(r"^(\d{4})", part)
        if m:
            year = int(m.group(1))
            if 2000 <= year <= 2100:
                return year
    return None


def years_in_window(cutoff_date: datetime) -> set[int]:
    today = datetime.today()
    years, y = set(), cutoff_date.year
    while y <= today.year:
        years.add(y)
        y += 1
    return years


def parse_date(date_str: str) -> datetime | None:
    for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%B %d, %Y"):
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None


# ── CAPTCHA ───────────────────────────────────────────────────────────────────

def captcha_present(driver) -> bool:
    try:
        driver.find_element(By.ID, CAPTCHA_IMG_ID)
        return True
    except NoSuchElementException:
        return False


def solve_captcha(driver, wait, model, device, chars) -> bool:
    """Solve whatever CAPTCHA is on screen. Returns True on success."""
    for attempt in range(1, MAX_CAPTCHA_TRIES + 1):
        if not captcha_present(driver):
            return True

        img_el = driver.find_element(By.ID, CAPTCHA_IMG_ID)
        img_src = img_el.get_attribute("src") or ""
        img_url = img_src if img_src.startswith("http") else BASE_URL + img_src

        if CAPTCHA_AUTO and model is not None:
            try:
                gray  = download_image(img_url)
                guess = predict(gray, model, device, chars)
            except Exception as e:
                print(f"    Image error: {e} — refreshing")
                driver.refresh()
                time.sleep(2)
                continue
        else:
            guess = input(f"  Enter CAPTCHA (attempt {attempt}): ").strip()

        # ── Debug: save the image named after the guess ───────────────────────
        try:
            import cv2 as _cv2, os as _os
            debug_dir = Path("captcha_debug")
            debug_dir.mkdir(exist_ok=True)
            # Include attempt number to avoid collisions when the same guess repeats
            debug_path = debug_dir / f"{guess}_{attempt:02d}.png"
            _cv2.imwrite(str(debug_path), gray)
        except Exception as _e:
            print(f"    Debug save failed: {_e}")

        print(f"  CAPTCHA attempt {attempt:2d}: '{guess}'")
        try:
            inp = driver.find_element(By.NAME, CAPTCHA_INPUT)
            inp.clear()
            inp.send_keys(guess)
            driver.find_element(By.NAME, CAPTCHA_SUBMIT).click()
        except NoSuchElementException as e:
            print(f"    Missing element: {e}")
            return False

        time.sleep(1.5)

        if not captcha_present(driver):
            print(f"  ✓ CAPTCHA solved on attempt {attempt}!")
            return True
        print("    Wrong — retrying...")

    print("  ✗ Could not solve CAPTCHA after max attempts.")
    return False


# ── Excel ─────────────────────────────────────────────────────────────────────

def load_or_create_workbook(path: Path):
    if path.exists():
        # Guard against a corrupted file (e.g. truncated mid-save crash).
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as e:
            # Rename the bad file so we don't lose it, then start fresh.
            bad_path = path.with_suffix(".corrupted.xlsx")
            path.rename(bad_path)
            print(f"  WARNING: could not open {path.name} ({e})")
            print(f"  Renamed corrupted file to {bad_path.name} — starting fresh.")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Cases"
            _write_header(ws)
            return wb, ws, set()

        ws = wb.active
        existing = set()
        for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
            if row[0]:
                existing.add(str(row[0]).strip())
        return wb, ws, existing

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cases"
    _write_header(ws)
    return wb, ws, set()


def _write_header(ws):
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
        cell.fill = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
    ws.row_dimensions[1].height = 18
    for i, w in enumerate([22, 45, 28, 16, 22], start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def append_case(ws, existing, case_num, title, court, filing_date, search_name) -> bool:
    if case_num in existing:
        print(f"    Duplicate — skipping {case_num}")
        return False
    row_num = ws.max_row + 1
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    fill_color = ALT_ROW_COLOR if row_num % 2 == 0 else "FFFFFF"
    for col_idx, value in enumerate([case_num, title, court, filing_date, search_name], start=1):
        cell = ws.cell(row=row_num, column=col_idx, value=value)
        cell.font = Font(name="Arial", size=10)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.alignment = Alignment(vertical="center")
        cell.border = border
    existing.add(case_num)
    print(f"    ✓ Added: {case_num}  |  {filing_date}  |  {title[:50]}")
    return True


# ── Page helpers ──────────────────────────────────────────────────────────────

def get_court_name(driver) -> str:
    try:
        links = driver.find_elements(By.XPATH, "//a[contains(@href,'courtinfo.aspx')]")
        if links:
            return links[0].text.strip()
    except Exception:
        pass
    try:
        el = driver.find_element(
            By.XPATH,
            "//*[contains(text(),'Court')]/following-sibling::*[1]"
        )
        return el.text.strip()
    except Exception:
        return ""


def on_search_page(driver) -> bool:
    """
    True only when on the bare search/CAPTCHA form with no results table.
    The results page shares the same caselookup URL, so URL alone is not enough.
    """
    if "caselookup" not in driver.current_url.lower():
        return False
    # If a results table is present we are on the results page, not the search form
    try:
        driver.find_element(By.XPATH, "//*[contains(@id,'gvSearchResults')]")
        return False  # results table present — this is the results page
    except NoSuchElementException:
        return True   # no results table — bare search/CAPTCHA form


def on_captcha_page(driver) -> bool:
    return captcha_present(driver)


def click_case_search_tab(driver) -> bool:
    """
    Click the link that returns to the case search page.
    Priority order:
      1. "Go To Case Search" — the exact link on the Page Error screen
      2. "Go to Case Search" — same link, alternate capitalisation
      3. "Case Search"       — the nav tab on normal pages
      4. Any href=caselookup.aspx fallback
    Returns True if a link was found and clicked.
    """
    # If we are on an error page, look for "Go To Case Search" by its exact text
    try:
        body_text = driver.find_element(By.TAG_NAME, "body").text
        if "Page Error" in body_text:
            # Use exact link text from the error page screenshot
            try:
                link = driver.find_element(
                    By.XPATH, "//a[normalize-space(.)='Go To Case Search']"
                )
                link.click()
                time.sleep(1.5)
                print("  ↩ Clicked 'Go To Case Search' on error page")
                return True
            except NoSuchElementException:
                pass
    except Exception:
        pass

    # General fallback: try text variants then any caselookup.aspx href
    for text_fragment in ("Go To Case Search", "Go to Case Search", "Case Search"):
        try:
            link = driver.find_element(
                By.XPATH,
                f"//a[contains(@href,'caselookup.aspx') and "
                f"contains(normalize-space(.),'{text_fragment}')]"
            )
            link.click()
            time.sleep(1.5)
            print(f"  ↩ Clicked '{text_fragment}' link — back to known page")
            return True
        except NoSuchElementException:
            continue

    # Broadest fallback: any caselookup.aspx link on the page
    try:
        link = driver.find_element(
            By.XPATH, "//a[contains(@href,'caselookup.aspx')]"
        )
        link.click()
        time.sleep(1.5)
        print("  ↩ Clicked caselookup.aspx link — back to known page")
        return True
    except NoSuchElementException:
        return False


def is_known_page(driver) -> bool:
    """
    Return True if the current page is one we know how to handle:
      - the search / CAPTCHA form  (caselookup.aspx)
      - a case detail page         (casedetail.aspx or similar)
      - the search results         (same caselookup URL, table present)
    Any other page (session timeout, error, 404, etc.) is 'unknown'.
    """
    url = driver.current_url.lower()
    known_fragments = ("caselookup", "casedetail", "casesummary", "publicaccess")
    if not any(f in url for f in known_fragments):
        return False
    # Extra guard: error pages sometimes stay on the right URL but show an error banner
    try:
        body = driver.find_element(By.TAG_NAME, "body").text.lower()
        error_phrases = (
            "page error",                             # exact heading on the error page
            "there has been an error in the page",   # body text on the error page
            "we're sorry",                            # greeting on the error page
            "an error has occurred",
            "runtime error",
            "server error",
            "object reference",
            "page not found",
            "404",
            "500",
        )
        if any(p in body for p in error_phrases):
            return False
    except Exception:
        pass
    return True


def recover_unknown_page(
    driver, wait, model, device, chars,
    biz_name: str, current_page: int
) -> bool:
    """
    Called whenever we land on an unrecognised or error page.
    Strategy:
      1. Try clicking the 'Case Search' tab.
      2. Solve any CAPTCHA that appears.
      3. Re-submit the search for biz_name and re-sort.
      4. Navigate back to current_page if > 1.
    Returns True if recovery succeeded, False if we could not get back.
    """
    print(f"  ⚠ Unknown/error page detected — attempting recovery via 'Case Search' tab")

    # Step 1: click the tab (or hard-navigate if the tab itself is missing)
    if not click_case_search_tab(driver):
        print("  'Case Search' tab not found — hard-navigating to SEARCH_URL")
        driver.get(SEARCH_URL)
        time.sleep(1.5)

    # Step 2: CAPTCHA
    if not solve_captcha(driver, wait, model, device, chars):
        print("  CAPTCHA failed during recovery")
        return False

    # Step 3: re-search
    if not submit_search(driver, wait, biz_name):
        return False
    if not solve_captcha(driver, wait, model, device, chars):
        return False
    sort_ok = sort_results_twice(driver, wait, model, device, chars)
    if not sort_ok:
        print("  Sort interrupted during unknown-page recovery — aborting")
        return False

    # Step 4: restore page position
    if current_page > 1:
        if not goto_page(driver, wait, current_page, model, device, chars):
            return False

    print(f"  ✓ Recovery complete — back on page {current_page} for {biz_name!r}")
    return True


def submit_search(driver, wait, biz_name: str) -> bool:
    """
    Fill txtLName with biz_name and submit. Returns True on successful submit.

    The AZ Courts form hides/shows panels based on a search-type radio button.
    We must click the correct radio FIRST so that txtLName becomes interactive,
    then fill it and submit.

    Known radio button name: ctl00$ContentPlaceHolder1$rbSearchType
    Known values observed on the form:
      "N" = Name (individual)   "B" = Business   "C" = Case number
    We try "B" first (business), then "N" as fallback, then just attempt the
    field directly in case the form has no tabs at all.
    """
    # ── Step 1: activate the correct search panel ─────────────────────────────
    # Try to click the Business ("B") radio; fall back to Name ("N"); then skip.
    activated = False
    for radio_value in ("B", "N"):
        try:
            radio = driver.find_element(
                By.XPATH,
                f"//input[@type='radio' and @name='ctl00$ContentPlaceHolder1$rbSearchType'"
                f" and @value='{radio_value}']"
            )
            driver.execute_script("arguments[0].click();", radio)
            time.sleep(0.6)   # allow panel to show/hide via JS
            activated = True
            print(f"  Selected search tab: value='{radio_value}'")
            break
        except NoSuchElementException:
            continue

    if not activated:
        print("  No search-type radio found — attempting field directly")

    # ── Step 2: locate the name field (may be inside a now-visible panel) ────
    # Try the confirmed name first, then common ASP.NET alternatives.
    field_names = [
        "ctl00$ContentPlaceHolder1$txtLName",
        "ctl00$ContentPlaceHolder1$txtBusinessName",
        "ctl00$ContentPlaceHolder1$txtLastName",
    ]
    inp = None
    for fname in field_names:
        try:
            candidate = driver.find_element(By.NAME, fname)
            # Make sure it is actually visible / enabled
            if candidate.is_displayed() and candidate.is_enabled():
                inp = candidate
                print(f"  Using field: {fname}")
                break
            else:
                # Try JS-forced interaction as last resort for this field name
                inp = candidate
                print(f"  Field {fname} found but not displayed — will use JS")
                break
        except NoSuchElementException:
            continue

    if inp is None:
        # Dump all visible inputs so we can see exactly what the page has
        all_inputs = driver.find_elements(By.XPATH, "//input[@type='text' or @type='search']")
        names = [(el.get_attribute("name"), el.get_attribute("id"),
                  el.is_displayed(), el.is_enabled()) for el in all_inputs]
        print(f"  ERROR: could not find any name/business field.")
        print(f"  Visible text inputs on page: {names}")
        return False

    # ── Step 3: fill the field ────────────────────────────────────────────────
    try:
        inp.clear()
        inp.send_keys(biz_name)
    except Exception:
        # Field may need JS interaction if it is hidden behind a panel
        driver.execute_script("arguments[0].value = arguments[1];", inp, biz_name)

    # ── Step 4: submit ────────────────────────────────────────────────────────
    submitted = False
    for btn_locator in [
        (By.XPATH, "//input[@type='submit' and contains(@name,'btnSearch')]"),
        (By.XPATH, "//input[@type='submit' and contains(@value,'Search')]"),
        (By.XPATH, "//input[@type='button' and contains(@value,'Search')]"),
        (By.XPATH, "//a[contains(text(),'Search')]"),
    ]:
        try:
            btn = driver.find_element(*btn_locator)
            driver.execute_script("arguments[0].click();", btn)
            submitted = True
            break
        except NoSuchElementException:
            continue

    if not submitted:
        print("  No Search button found — pressing Enter")
        try:
            inp.send_keys(Keys.RETURN)
        except Exception:
            driver.execute_script(
                "arguments[0].dispatchEvent(new KeyboardEvent('keypress',{key:'Enter',keyCode:13,bubbles:true}));",
                inp
            )

    time.sleep(2)
    return True


def sort_results_twice(driver, wait, model, device, chars) -> bool:
    """
    Fire the sort postback twice so results end up newest-first.
    Returns True  if both clicks completed and we are still on the results page.
    Returns False if a CAPTCHA redirect sent us back to the search page mid-sort
                  (caller must re-do the full search+sort sequence).
    """
    for click_num in range(1, 3):
        # Wait for results table before firing JS
        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(@id,'gvSearchResults')]")
            ))
        except TimeoutException:
            print(f"  Sort click {click_num}: results table not found — aborting sort")
            return False

        try:
            driver.execute_script(SORT_JS)
        except Exception as e:
            print(f"  Sort click {click_num} execute_script failed: {e} — aborting sort")
            return False

        print(f"  Sort click {click_num}/2")
        time.sleep(1.5)

        # CAPTCHA may appear after the postback
        if captcha_present(driver):
            if not solve_captcha(driver, wait, model, device, chars):
                print(f"  CAPTCHA failed after sort click {click_num}")
                return False
            # Solved but now on search page — sort was interrupted
            if on_search_page(driver):
                print(f"  Sort click {click_num} triggered CAPTCHA redirect to search page")
                return False

    return True


def goto_page(driver, wait, page_num: int, model, device, chars) -> bool:
    """
    Navigate to a specific results page by clicking the pager anchor whose href
    contains Page$N — exactly what the browser does when a user clicks the link.
    Falls back to execute_script if the link is not visible in the DOM.
    Returns True  if the results table is present after navigation.
    Returns False on failure or CAPTCHA redirect.
    """
    if page_num == 1:
        return True  # already there after sort

    # Wait for the results table to be fully settled before navigating
    try:
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//*[contains(@id,'gvSearchResults')]")
        ))
    except TimeoutException:
        print(f"  goto_page: results table not ready before page {page_num} navigation")
        return False

    # Strategy 1: find and click the actual pager link in the DOM.
    # The href looks like: javascript:__doPostBack('...','Page$2')
    pager_xpath = "//a[contains(@href,'Page$" + str(page_num) + "')]"
    try:
        pager_link = driver.find_element(By.XPATH, pager_xpath)
        driver.execute_script("arguments[0].click();", pager_link)
        print(f"  Clicked pager link for page {page_num}")
    except NoSuchElementException:
        # Strategy 2: fire the postback directly via JS
        print(f"  Pager link for page {page_num} not in DOM — using execute_script")
        js = "__doPostBack('" + PAGE_JS + "','Page$" + str(page_num) + "')"
        try:
            driver.execute_script(js)
        except Exception as e:
            print(f"  goto_page execute_script failed for page {page_num}: {e}")
            return False

    time.sleep(1.0)

    # If a CAPTCHA appeared after the page click, solve it here.
    # After solving, check whether we are back on the results table (success)
    # or on the bare search form (redirect — signal caller to recover).
    if captcha_present(driver):
        if not solve_captcha(driver, wait, model, device, chars):
            return False
        # Check what page we landed on after solving
        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(@id,'gvSearchResults')]")
            ))
            return True  # CAPTCHA solved and results table is present
        except TimeoutException:
            # Solved but no results table — we were redirected to search form
            print(f"  goto_page: CAPTCHA solved but no results table — redirected")
            return False

    # No CAPTCHA — check directly for results table
    try:
        wait.until(EC.presence_of_element_located(
            (By.XPATH, "//*[contains(@id,'gvSearchResults')]")
        ))
        return True
    except TimeoutException:
        print(f"  goto_page: results table not found after navigating to page {page_num}")
        return False


def read_page_case_links(driver) -> list[tuple[str, str]]:
    """
    Return (element_id, case_number_text) for every case-number link on the
    current results page.
    """
    links = driver.find_elements(
        By.XPATH,
        "//*[contains(@id,'gvSearchResults') and contains(@id,'lbCaseNum')]"
    )
    return [(el.get_attribute("id"), el.text.strip()) for el in links]


def has_next_page(driver, current_page: int) -> bool:
    """
    Check whether there is a next page.
    ASP.NET gridview pagers only render a window of page numbers so the next
    page link may not appear in the DOM even when more pages exist.  We use two
    strategies:
      1. Look for the explicit next-page link/number in the pager (fast).
      2. Fallback: check that the current page is not the last — if the pager
         contains any page link at all, assume more pages may follow and let
         goto_page determine the truth.
    """
    next_num = current_page + 1
    # Strategy 1: explicit link for next page number
    try:
        driver.find_element(
            By.XPATH,
            f"//a[contains(@href,\"Page${next_num}\") or normalize-space(text())='{next_num}']"
        )
        return True
    except NoSuchElementException:
        pass
    # Strategy 2: pager exists at all → more pages may follow; let goto_page confirm
    try:
        driver.find_element(
            By.XPATH,
            "//*[contains(@id,'gvSearchResults')]//a[contains(@href,'Page$')]"
        )
        return True
    except NoSuchElementException:
        return False


# ── Core: process one business name ──────────────────────────────────────────

def process_business_name(
    biz_name: str,
    driver, wait,
    model, device, chars,
    valid_years: set[int],
    cutoff: datetime,
    wb, ws, existing_cases: set,
    out_path: Path,
    new_count_ref: list,   # [int] — mutable counter passed by ref
):
    """
    Search for biz_name, sort results, then page through ALL pages collecting
    every case whose case-number year is in valid_years.  For each such case,
    open it, check the filing date, and record it if within the cutoff window.

    Stops moving to a new page / new name only when a case-number year is
    found that is NOT in valid_years (i.e. too old).

    If a mid-search CAPTCHA redirects us to the search page, we re-search the
    same name, re-sort, navigate back to the correct page, skip cases we have
    already seen (by case number), and continue.
    """

    print(f"\n{'='*60}")
    print(f"Searching: {biz_name!r}")
    print(f"{'='*60}")

    # ── Search + sort (retry loop) ───────────────────────────────────────────
    # If the sort triggers a CAPTCHA that redirects back to the search page,
    # we must re-do the full search+sort — loop until we land on sorted results.
    MAX_SEARCH_ATTEMPTS = 5
    for search_attempt in range(1, MAX_SEARCH_ATTEMPTS + 1):
        if search_attempt > 1:
            print(f"  Re-attempting search+sort (attempt {search_attempt})...")
            driver.get(SEARCH_URL)
            time.sleep(1.5)
            if not solve_captcha(driver, wait, model, device, chars):
                print("  CAPTCHA failed on re-attempt — skipping")
                return

        if not submit_search(driver, wait, biz_name):
            return

        if not solve_captcha(driver, wait, model, device, chars):
            print("  CAPTCHA failed after search — skipping")
            return

        # Check for no-results before attempting sort.
        # The site shows "No cases found matching your search criteria." as plain
        # text — check for it first so we never fire sort clicks against an empty page.
        body = driver.find_element(By.TAG_NAME, "body").text
        if "No cases found matching your search criteria" in body:
            print(f"  No results for {biz_name!r}")
            return

        # Wait for the results table to confirm results are present
        try:
            wait.until(EC.presence_of_element_located(
                (By.XPATH, "//*[contains(@id,'gvSearchResults')]")
            ))
        except TimeoutException:
            body_lower = body.lower()
            if any(p in body_lower for p in ["no cases", "no records", "no results", "not found"]):
                print(f"  No results for {biz_name!r}")
                return
            else:
                print(f"  Results table not found:\n{body[:400]}")
                return

        sort_ok = sort_results_twice(driver, wait, model, device, chars)
        if sort_ok:
            break  # successfully on sorted results — proceed to page loop
        # sort was interrupted; loop will re-search
    else:
        print(f"  Could not complete search+sort after {MAX_SEARCH_ATTEMPTS} attempts — skipping")
        return

    # ── Page loop ─────────────────────────────────────────────────────────────
    current_page = 1
    done_with_name = False          # set True when an out-of-range year is seen
    seen_case_nums: set[str] = set()  # tracks what we've already clicked this run

    while not done_with_name:
        print(f"\n  -- Page {current_page} --")

        # Read all case links on this page
        page_links = read_page_case_links(driver)
        print(f"  {len(page_links)} case link(s) on page {current_page}")

        if not page_links:
            break

        # ── Check each link on this page ──────────────────────────────────────
        for link_id, case_text in page_links:
            year = extract_year_from_case(case_text)

            # year=None means the case number format is unrecognised — skip it,
            # do NOT stop: there may be valid-year cases further down the page.
            if year is None:
                print(f"  Could not parse year from '{case_text}' — skipping")
                continue

            # Out-of-range year → we have passed all relevant cases, stop this name
            if year not in valid_years:
                print(f"  Year {year} not in {sorted(valid_years)} "
                      f"({case_text}) — done with {biz_name!r}")
                done_with_name = True
                break

            # Already processed this case in a prior CAPTCHA-recovery iteration
            if case_text in seen_case_nums:
                print(f"  Already processed {case_text} — skipping")
                continue

            # Duplicate check against Excel — case_text IS the case number,
            # no need to click in just to read it again.
            if case_text in existing_cases:
                print(f"    Duplicate — skipping {case_text}")
                seen_case_nums.add(case_text)
                continue

            # ── Click into the case ───────────────────────────────────────────
            print(f"\n  → Opening case: {case_text}")
            try:
                link_el = driver.find_element(By.ID, link_id)
                link_el.click()
                time.sleep(1.5)
            except (NoSuchElementException, StaleElementReferenceException) as e:
                print(f"    Could not click {link_id}: {e}")
                seen_case_nums.add(case_text)
                continue

            # ── Mid-click CAPTCHA check ───────────────────────────────────────
            # If a CAPTCHA appears AND solving it returns us to the search page,
            # we must re-search and navigate back to where we were.
            if on_captcha_page(driver):
                print("  ⚠ CAPTCHA triggered mid-search — solving and resuming...")
                if not solve_captcha(driver, wait, model, device, chars):
                    print("  CAPTCHA failed — aborting this name")
                    return

                # After solving, are we back on the search page?
                if on_search_page(driver):
                    print(f"  Redirected to search page — re-searching {biz_name!r} "
                          f"and returning to page {current_page}...")
                    driver.get(SEARCH_URL)
                    time.sleep(1.5)
                    solve_captcha(driver, wait, model, device, chars)
                    if not submit_search(driver, wait, biz_name):
                        return
                    if not solve_captcha(driver, wait, model, device, chars):
                        return
                    sort_ok = sort_results_twice(driver, wait, model, device, chars)
                    if not sort_ok:
                        print("  Sort interrupted during mid-click recovery — aborting")
                        return
                    if current_page > 1:
                        if not goto_page(driver, wait, current_page, model, device, chars):
                            return
                    # Do NOT add case_text to seen_case_nums — we want to re-check
                    # this case after recovery in case we missed something.
                    break

            # ── Read filing date ──────────────────────────────────────────────
            filing_date_str = ""
            try:
                fd_el = driver.find_element(
                    By.ID, "ctl00_ContentPlaceHolder1_gvCaseInfo_ctl02_tcFileDate"
                )
                filing_date_str = fd_el.text.strip()
            except NoSuchElementException:
                try:
                    fd_el = driver.find_element(
                        By.XPATH,
                        "//*[contains(text(),'Filed') or contains(text(),'Filing')]"
                        "/following-sibling::*[1]"
                    )
                    filing_date_str = fd_el.text.strip()
                except NoSuchElementException:
                    print("    Could not find filing date — skipping")
                    seen_case_nums.add(case_text)
                    driver.back()
                    time.sleep(1)
                    if not is_known_page(driver):
                        if not recover_unknown_page(
                            driver, wait, model, device, chars,
                            biz_name, current_page
                        ):
                            return
                        break  # re-read page links after recovery
                    solve_captcha(driver, wait, model, device, chars)
                    continue

            filing_dt = parse_date(filing_date_str)
            if filing_dt is None:
                print(f"    Unparseable date '{filing_date_str}' — skipping")
                seen_case_nums.add(case_text)
                driver.back()
                time.sleep(1)
                if not is_known_page(driver):
                    if not recover_unknown_page(
                        driver, wait, model, device, chars,
                        biz_name, current_page
                    ):
                        return
                    break  # re-read page links after recovery
                solve_captcha(driver, wait, model, device, chars)
                continue

            # Date is within our window — collect the case data
            if filing_dt >= cutoff:
                # Use the case number we already have from the results table.
                case_num = case_text

                title = ""
                try:
                    t_el = driver.find_element(
                        By.ID, "ctl00_ContentPlaceHolder1_gvCaseInfo_ctl02_tcTitle"
                    )
                    title = t_el.text.strip()
                except NoSuchElementException:
                    try:
                        t_el = driver.find_element(
                            By.XPATH,
                            "//*[contains(text(),'Title') or contains(text(),'Case Name')]"
                            "/following-sibling::*[1]"
                        )
                        title = t_el.text.strip()
                    except NoSuchElementException:
                        pass

                court = get_court_name(driver)

                added = append_case(
                    ws, existing_cases,
                    case_num, title, court, filing_date_str, biz_name
                )
                if added:
                    new_count_ref[0] += 1
                    wb.save(out_path)
            else:
                # First out-of-range filing date — cases are newest-first so
                # nothing further can be in range. Stop this business name.
                print(f"    Filing date {filing_date_str} out of range — done with {biz_name!r}")
                done_with_name = True

            # Go back to results
            driver.back()
            time.sleep(1)

            # Unknown / error page → recover via Case Search tab
            if not is_known_page(driver):
                if not recover_unknown_page(
                    driver, wait, model, device, chars,
                    biz_name, current_page
                ):
                    return
                # Do NOT mark as seen — re-check this case after recovery.
                break  # re-read page links after recovery

            # CAPTCHA check on return
            if on_captcha_page(driver):
                print("  ⚠ CAPTCHA on return — solving and resuming...")
                if not solve_captcha(driver, wait, model, device, chars):
                    return
                if on_search_page(driver):
                    print(f"  Redirected to search — re-searching {biz_name!r}, "
                          f"page {current_page}...")
                    driver.get(SEARCH_URL)
                    time.sleep(1.5)
                    solve_captcha(driver, wait, model, device, chars)
                    if not submit_search(driver, wait, biz_name):
                        return
                    if not solve_captcha(driver, wait, model, device, chars):
                        return
                    sort_ok = sort_results_twice(driver, wait, model, device, chars)
                    if not sort_ok:
                        print("  Sort interrupted during return recovery — aborting")
                        return
                    if current_page > 1:
                        if not goto_page(driver, wait, current_page, model, device, chars):
                            return
                    # Do NOT mark as seen — re-check this case after recovery.
                    break  # re-read page links

            # Only mark as seen once we have cleanly returned to the results page.
            seen_case_nums.add(case_text)

        else:
            # Inner for-loop completed without a break → try next page
            if done_with_name:
                break
            if has_next_page(driver, current_page):
                current_page += 1
                print(f"\n  → Advancing to page {current_page}")
                page_ok = goto_page(driver, wait, current_page, model, device, chars)

                if not page_ok:
                    # goto_page returns False for two reasons:
                    # (a) CAPTCHA redirect → recover and retry the page
                    # (b) No more pages    → stop
                    if on_search_page(driver) or captcha_present(driver):
                        print(f"  Redirected to search — re-searching and going to page {current_page}...")
                        driver.get(SEARCH_URL)
                        time.sleep(1.5)
                        if not solve_captcha(driver, wait, model, device, chars):
                            return
                        if not submit_search(driver, wait, biz_name):
                            return
                        if not solve_captcha(driver, wait, model, device, chars):
                            return
                        sort_ok = sort_results_twice(driver, wait, model, device, chars)
                        if not sort_ok:
                            print(f"  Sort interrupted during page recovery — retrying from top")
                            return  # let the outer search-attempt loop retry
                        if not goto_page(driver, wait, current_page, model, device, chars):
                            return
                    else:
                        print(f"  No more pages for {biz_name!r} (page {current_page} not found)")
                        break
            else:
                print(f"  No more pages for {biz_name!r}")
                break
            continue  # re-enter while loop for new page

        # A break inside the inner for-loop means we need to re-read page links
        # (CAPTCHA recovery mid-page) — continue the while loop without advancing
        if done_with_name:
            break
        # else: fall through to re-read the current page (seen_case_nums guards dups)


# ── Entry point ───────────────────────────────────────────────────────────────

def scrape(out_path: Path, days: int, headless: bool):
    cutoff = datetime.today() - timedelta(days=days)
    valid_years = years_in_window(cutoff)
    print(f"Date window: {cutoff.strftime('%m/%d/%Y')} – today  |  Years: {sorted(valid_years)}")

    model = device = chars = None
    if CAPTCHA_AUTO:
        model, device, chars = load_model()

    opts = webdriver.ChromeOptions()
    if headless:
        opts.add_argument("--headless=new")
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=opts)
    wait = WebDriverWait(driver, 15)

    wb, ws, existing_cases = load_or_create_workbook(out_path)
    new_count_ref = [0]  # mutable int passed into helper

    try:
        driver.get(SEARCH_URL)

        if not solve_captcha(driver, wait, model, device, chars):
            print("Could not pass initial CAPTCHA — aborting.")
            return

        for biz_name in BUSINESS_NAMES:
            process_business_name(
                biz_name, driver, wait,
                model, device, chars,
                valid_years, cutoff,
                wb, ws, existing_cases,
                out_path, new_count_ref,
            )

            # Always hard-navigate to the search URL between names so the form
            # is freshly rendered — the results page shares the same URL prefix
            # so on_search_page() alone is not a reliable guard.
            driver.get(SEARCH_URL)
            time.sleep(1.5)
            solve_captcha(driver, wait, model, device, chars)

    finally:
        wb.save(out_path)
        driver.quit()
        print(f"\n{'='*60}")
        print(f"Done. {new_count_ref[0]} new case(s) added to {out_path}")
        print(f"{'='*60}")


def main():
    parser = argparse.ArgumentParser(description="AZ Courts JG Wentworth Scraper")
    parser.add_argument("--out",      default="AZcases_Last60.xlsx")
    parser.add_argument("--days",     type=int, default=60)
    parser.add_argument("--headless", action="store_true")
    args = parser.parse_args()
    scrape(Path(args.out), args.days, args.headless)


if __name__ == "__main__":
    main()